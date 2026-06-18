"""
Import Guangdong 2023 Input-Output Table (42-sector) into SQLite.

Usage (from backend/):
    python -m scripts.init_io_table

Reads: 0331基础资料补充/广东投入产出比.xlsx (relative to project root)
Writes: companies.db — tables io_sectors, io_flow_matrix, io_final_use, io_industry_mapping
"""

from __future__ import annotations

import sys
from pathlib import Path

from sqlalchemy import Float, ForeignKey, Index, Integer, String, Text, UniqueConstraint, inspect
from sqlalchemy.orm import Mapped, mapped_column

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

_BACKEND_DIR = Path(__file__).resolve().parents[1]
_IO_XLSX = _BACKEND_DIR.parent.parent.parent.parent.joinpath("0331基础资料补充", "广东投入产出比.xlsx")

# ---------------------------------------------------------------------------
# Industry mapping: company industry name → IO sector code
# ---------------------------------------------------------------------------

_INDUSTRY_MAPPING: list[tuple[str, str, str, str]] = [
    # (company_industry, io_code, confidence, note)
    # --- 第一产业 → 01 ---
    ("农业", "01", "exact", ""),
    ("林业", "01", "exact", ""),
    ("畜牧业", "01", "exact", ""),
    ("渔业", "01", "exact", ""),
    ("农、林、牧、渔专业及辅助性活动", "01", "exact", ""),
    # --- 采矿业 → 02-05 ---
    ("煤炭开采和洗选业", "02", "exact", ""),
    ("黑色金属矿采选业", "04", "exact", "黑色金属→金属矿"),
    ("有色金属矿采选业", "04", "exact", "有色金属→金属矿"),
    ("非金属矿采选业", "05", "exact", ""),
    ("其他采矿业", "05", "broad", ""),
    ("开采专业及辅助性活动", "04", "broad", "采矿辅助→采矿"),
    # --- 制造业 → 06-23 ---
    ("农副食品加工业", "06", "exact", ""),
    ("食品制造业", "06", "exact", ""),
    ("酒、饮料和精制茶制造业", "06", "exact", ""),
    ("烟草制品业", "06", "exact", ""),
    ("纺织业", "07", "exact", ""),
    ("纺织服装、服饰业", "08", "exact", ""),
    ("皮革、毛皮、羽毛及其制品和制鞋业", "08", "exact", ""),
    ("木材加工和木、竹、藤、棕、草制品业", "09", "exact", ""),
    ("家具制造业", "09", "exact", ""),
    ("造纸和纸制品业", "10", "exact", ""),
    ("印刷和记录媒介复制业", "10", "exact", ""),
    ("文教、工美、体育和娱乐用品制造业", "10", "exact", ""),
    ("石油、煤炭及其他燃料加工业", "11", "exact", ""),
    ("化学原料和化学制品制造业", "12", "exact", ""),
    ("化学纤维制造业", "12", "exact", ""),
    ("医药制造业", "12", "broad", "医药属于化学工业分支"),
    ("橡胶和塑料制品业", "12", "broad", "橡塑属于化学相关"),
    ("非金属矿物制品业", "13", "exact", ""),
    ("黑色金属冶炼和压延加工业", "14", "exact", ""),
    ("有色金属冶炼和压延加工业", "14", "exact", ""),
    ("金属制品业", "15", "exact", ""),
    ("通用设备制造业", "16", "exact", ""),
    ("专用设备制造业", "17", "exact", ""),
    ("汽车制造业", "18", "exact", ""),
    ("铁路、船舶、航空航天和其他运输设备制造业", "18", "exact", ""),
    ("电气机械和器材制造业", "19", "exact", ""),
    ("计算机、通信和其他电子设备制造业", "20", "exact", ""),
    ("仪器仪表制造业", "21", "exact", ""),
    ("废弃资源综合利用业", "22", "exact", ""),
    ("其他制造业", "22", "exact", ""),
    ("金属制品、机械和设备修理业", "23", "exact", ""),
    # --- 公用事业与建筑 → 24-27 ---
    ("电力、热力生产和供应业", "24", "exact", ""),
    ("燃气生产和供应业", "25", "exact", ""),
    ("水的生产和供应业", "26", "exact", ""),
    ("房屋建筑业", "27", "exact", ""),
    ("土木工程建筑业", "27", "exact", ""),
    ("建筑安装业", "27", "exact", ""),
    ("建筑装饰、装修和其他建筑业", "27", "exact", ""),
    # --- 第三产业 → 28-42 ---
    ("批发业", "28", "exact", ""),
    ("零售业", "28", "exact", ""),
    ("道路运输业", "29", "exact", ""),
    ("水上运输业", "29", "exact", ""),
    ("铁路运输业", "29", "exact", ""),
    ("航空运输业", "29", "exact", ""),
    ("装卸搬运和仓储业", "29", "exact", ""),
    ("邮政业", "29", "exact", ""),
    ("多式联运和运输代理业", "29", "exact", ""),
    ("住宿业", "30", "exact", ""),
    ("餐饮业", "30", "exact", ""),
    ("电信、广播电视和卫星传输服务", "31", "exact", ""),
    ("互联网和相关服务", "31", "exact", ""),
    ("软件和信息技术服务业", "31", "exact", ""),
    ("货币金融服务", "32", "exact", ""),
    ("资本市场服务", "32", "exact", ""),
    ("保险业", "32", "exact", ""),
    ("其他金融业", "32", "exact", ""),
    ("房地产业", "33", "exact", ""),
    ("租赁业", "34", "exact", ""),
    ("商务服务业", "34", "exact", ""),
    ("研究和试验发展", "35", "exact", ""),
    ("专业技术服务业", "36", "exact", ""),
    ("科技推广和应用服务业", "36", "exact", ""),
    ("水利管理业", "37", "exact", ""),
    ("生态保护和环境治理业", "37", "exact", ""),
    ("公共设施管理业", "37", "exact", ""),
    ("居民服务业", "38", "exact", ""),
    ("机动车、电子产品和日用产品修理业", "38", "exact", ""),
    ("其他服务业", "38", "exact", ""),
    ("教育", "39", "exact", ""),
    ("卫生", "40", "exact", ""),
    ("社会工作", "40", "exact", ""),
    ("新闻和出版业", "41", "exact", ""),
    ("广播、电视、电影和录音制作业", "41", "exact", ""),
    ("文化艺术业", "41", "exact", ""),
    ("体育", "41", "exact", ""),
    ("娱乐业", "41", "exact", ""),
    ("土地管理业", "42", "broad", "归入公共管理"),
]

# ---------------------------------------------------------------------------
# Sector category lookup
# ---------------------------------------------------------------------------

_SECTOR_CATEGORY = {
    "01": "1",
    **{str(i): "2" for i in range(2, 28)},
    **{str(i): "3" for i in range(28, 43)},
}

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main() -> None:
    import openpyxl
    from app.core.database import Base, company_engine
    from sqlalchemy import select as sa_select
    from sqlalchemy.orm import Session

    # --- ensure openpyxl available ---
    try:
        import openpyxl as _  # noqa: F811
    except ImportError:
        print("openpyxl is required: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    # --- define models inline (no other file changes needed) ---

    class IoSector(Base):
        __tablename__ = "io_sectors"

        id: Mapped[int] = mapped_column(Integer, primary_key=True, autoincrement=True)
        code: Mapped[str] = mapped_column(String(4), unique=True, nullable=False)
        name: Mapped[str] = mapped_column(String(128), nullable=False)
        category: Mapped[str] = mapped_column(String(4), nullable=False)
        total_output: Mapped[float | None] = mapped_column(Float, nullable=True)
        total_input: Mapped[float | None] = mapped_column(Float, nullable=True)
        intermediate_input_total: Mapped[float | None] = mapped_column(Float, nullable=True)
        value_added_total: Mapped[float | None] = mapped_column(Float, nullable=True)
        labor_compensation: Mapped[float | None] = mapped_column(Float, nullable=True)
        net_tax_production: Mapped[float | None] = mapped_column(Float, nullable=True)
        fixed_asset_depreciation: Mapped[float | None] = mapped_column(Float, nullable=True)
        operating_surplus: Mapped[float | None] = mapped_column(Float, nullable=True)

    class IoFlowMatrix(Base):
        __tablename__ = "io_flow_matrix"
        __table_args__ = (
            UniqueConstraint("row_sector_id", "col_sector_id", name="uq_flow_rc"),
        )

        id: Mapped[int] = mapped_column(Integer, primary_key=True, autoincrement=True)
        row_sector_id: Mapped[int] = mapped_column(Integer, ForeignKey("io_sectors.id"), nullable=False, index=True)
        col_sector_id: Mapped[int] = mapped_column(Integer, ForeignKey("io_sectors.id"), nullable=False, index=True)
        flow_value: Mapped[float] = mapped_column(Float, nullable=False)
        direct_consumption_coeff: Mapped[float | None] = mapped_column(Float, nullable=True)

    class IoFinalUse(Base):
        __tablename__ = "io_final_use"

        id: Mapped[int] = mapped_column(Integer, primary_key=True, autoincrement=True)
        sector_id: Mapped[int] = mapped_column(Integer, ForeignKey("io_sectors.id"), nullable=False, unique=True)
        rural_consumption: Mapped[float | None] = mapped_column(Float, nullable=True)
        urban_consumption: Mapped[float | None] = mapped_column(Float, nullable=True)
        household_consumption: Mapped[float | None] = mapped_column(Float, nullable=True)
        government_consumption: Mapped[float | None] = mapped_column(Float, nullable=True)
        total_consumption: Mapped[float | None] = mapped_column(Float, nullable=True)
        fixed_capital_formation: Mapped[float | None] = mapped_column(Float, nullable=True)
        inventory_change: Mapped[float | None] = mapped_column(Float, nullable=True)
        total_capital_formation: Mapped[float | None] = mapped_column(Float, nullable=True)
        exports: Mapped[float | None] = mapped_column(Float, nullable=True)
        inter_provincial_outflow: Mapped[float | None] = mapped_column(Float, nullable=True)
        total_final_use: Mapped[float | None] = mapped_column(Float, nullable=True)
        imports: Mapped[float | None] = mapped_column(Float, nullable=True)
        inter_provincial_inflow: Mapped[float | None] = mapped_column(Float, nullable=True)

    class IoIndustryMapping(Base):
        __tablename__ = "io_industry_mapping"
        __table_args__ = (
            Index("ix_io_industry_mapping_company_industry", "company_industry"),
        )

        id: Mapped[int] = mapped_column(Integer, primary_key=True, autoincrement=True)
        company_industry: Mapped[str] = mapped_column(String(128), unique=True, nullable=False)
        io_sector_id: Mapped[int] = mapped_column(Integer, ForeignKey("io_sectors.id"), nullable=False)
        confidence: Mapped[str] = mapped_column(String(8), nullable=False, default="exact")
        note: Mapped[str | None] = mapped_column(Text, nullable=True)

    # --- create tables ---
    inspector = inspect(company_engine)
    for model in (IoSector, IoFlowMatrix, IoFinalUse, IoIndustryMapping):
        if not inspector.has_table(model.__tablename__):
            model.__table__.create(bind=company_engine)
            print(f"  created table: {model.__tablename__}")

    # --- check existing data ---
    with Session(company_engine) as db:
        existing = db.scalar(sa_select(IoSector.id).limit(1))
        if existing is not None:
            print("io_sectors already initialized, skipping import")
            return

    # --- read Excel ---
    if not _IO_XLSX.exists():
        print(f"file not found: {_IO_XLSX}", file=sys.stderr)
        sys.exit(1)

    print(f"reading {_IO_XLSX}", flush=True)
    wb = openpyxl.load_workbook(str(_IO_XLSX), data_only=True)
    ws = wb["42IO"]

    # column layout:
    #   Col D(4) = sector 01, Col E(5) = sector 02, ... Col AT(46) = sector 42
    #   Col AU(47) = rural_consumption ... Col BE(57) = total_final_use
    #   Col BF(58) = imports, Col BG(59) = inter_provincial_inflow, Col BH(60) = total_output
    SECTOR_COL_START = 4   # D
    SECTOR_COL_END = 45    # AS (42 sectors, codes 01-42; col AT=TIU is aggregate)
    TOTAL_OUTPUT_COL = 60  # BH

    # Row layout:
    #   Row 5 = sector names (header), Row 6 = sector codes
    #   Row 7-48 = intermediate flow (42 sectors), Row 49 = intermediate input total
    #   Row 50 = labor_compensation, 51 = net_tax, 52 = depreciation, 53 = operating_surplus
    #   Row 54 = value_added_total, Row 55 = total_input
    SECTOR_ROW_START = 7
    SECTOR_ROW_END = 48
    ROW_INTERMEDIATE_TOTAL = 49
    ROW_LABOR = 50
    ROW_NET_TAX = 51
    ROW_DEPRECIATION = 52
    ROW_SURPLUS = 53
    ROW_VALUE_ADDED_TOTAL = 54
    ROW_TOTAL_INPUT = 55

    # --- parse sectors ---
    sector_names: dict[int, str] = {}
    sector_codes: dict[int, str] = {}
    for col_idx in range(SECTOR_COL_START, SECTOR_COL_END + 1):
        name = ws.cell(row=5, column=col_idx).value
        code = ws.cell(row=6, column=col_idx).value
        if name and code:
            sector_names[col_idx] = str(name).strip()
            sector_codes[col_idx] = str(code).strip()

    # --- insert io_sectors ---
    with Session(company_engine) as db:
        sector_id_map: dict[int, int] = {}  # col_idx → db id
        for col_idx in sorted(sector_codes.keys()):
            code = sector_codes[col_idx]
            name = sector_names[col_idx]
            total_input = ws.cell(row=ROW_TOTAL_INPUT, column=col_idx).value
            total_output = ws.cell(row=TOTAL_OUTPUT_COL, column=row_idx_to_sector_row(col_idx)).value if False else None
            intermediate_total = ws.cell(row=ROW_INTERMEDIATE_TOTAL, column=col_idx).value
            value_added_total = ws.cell(row=ROW_VALUE_ADDED_TOTAL, column=col_idx).value
            labor = ws.cell(row=ROW_LABOR, column=col_idx).value
            net_tax = ws.cell(row=ROW_NET_TAX, column=col_idx).value
            depreciation = ws.cell(row=ROW_DEPRECIATION, column=col_idx).value
            surplus = ws.cell(row=ROW_SURPLUS, column=col_idx).value
            # total_output is at row corresponding to this sector in column BH
            total_output = ws.cell(row=SECTOR_ROW_START + (col_idx - SECTOR_COL_START), column=TOTAL_OUTPUT_COL).value

            sector = IoSector(
                code=code,
                name=name,
                category=_SECTOR_CATEGORY.get(code, "2"),
                total_output=_to_float(total_output),
                total_input=_to_float(total_input),
                intermediate_input_total=_to_float(intermediate_total),
                value_added_total=_to_float(value_added_total),
                labor_compensation=_to_float(labor),
                net_tax_production=_to_float(net_tax),
                fixed_asset_depreciation=_to_float(depreciation),
                operating_surplus=_to_float(surplus),
            )
            db.add(sector)
            db.flush()
            sector_id_map[col_idx] = sector.id

        db.commit()
        print(f"  inserted {len(sector_id_map)} sectors")

        # --- insert io_flow_matrix (42×42) ---
        flow_count = 0
        for row_idx in range(SECTOR_ROW_START, SECTOR_ROW_END + 1):
            row_col_idx = SECTOR_COL_START + (row_idx - SECTOR_ROW_START)
            row_sector_id = sector_id_map.get(row_col_idx)
            if row_sector_id is None:
                continue
            for col_idx in range(SECTOR_COL_START, SECTOR_COL_END + 1):
                col_sector_id = sector_id_map.get(col_idx)
                if col_sector_id is None:
                    continue
                val = ws.cell(row=row_idx, column=col_idx).value
                db.add(IoFlowMatrix(
                    row_sector_id=row_sector_id,
                    col_sector_id=col_sector_id,
                    flow_value=_to_float(val) or 0.0,
                    direct_consumption_coeff=None,
                ))
                flow_count += 1
        db.commit()
        print(f"  inserted {flow_count} flow matrix cells")

        # --- insert io_final_use (42 rows) ---
        fu_count = 0
        for row_idx in range(SECTOR_ROW_START, SECTOR_ROW_END + 1):
            row_col_idx = SECTOR_COL_START + (row_idx - SECTOR_ROW_START)
            sector_id = sector_id_map.get(row_col_idx)
            if sector_id is None:
                continue
            db.add(IoFinalUse(
                sector_id=sector_id,
                rural_consumption=_to_float(ws.cell(row=row_idx, column=47).value),
                urban_consumption=_to_float(ws.cell(row=row_idx, column=48).value),
                household_consumption=_to_float(ws.cell(row=row_idx, column=49).value),
                government_consumption=_to_float(ws.cell(row=row_idx, column=50).value),
                total_consumption=_to_float(ws.cell(row=row_idx, column=51).value),
                fixed_capital_formation=_to_float(ws.cell(row=row_idx, column=52).value),
                inventory_change=_to_float(ws.cell(row=row_idx, column=53).value),
                total_capital_formation=_to_float(ws.cell(row=row_idx, column=54).value),
                exports=_to_float(ws.cell(row=row_idx, column=55).value),
                inter_provincial_outflow=_to_float(ws.cell(row=row_idx, column=56).value),
                total_final_use=_to_float(ws.cell(row=row_idx, column=57).value),
                imports=_to_float(ws.cell(row=row_idx, column=58).value),
                inter_provincial_inflow=_to_float(ws.cell(row=row_idx, column=59).value),
            ))
            fu_count += 1
        db.commit()
        print(f"  inserted {fu_count} final use rows")

        # --- compute direct consumption coefficients ---
        # a_ij = X_ij / X_j (flow from i to j / total_input of j)
        _compute_direct_consumption_coefficients(db, IoFlowMatrix, IoSector, sector_id_map)

        # --- insert io_industry_mapping ---
        code_to_id = {sector_codes[c]: sid for c, sid in sector_id_map.items()}
        mapping_count = 0
        for company_industry, io_code, confidence, note in _INDUSTRY_MAPPING:
            io_sector_id = code_to_id.get(io_code)
            if io_sector_id is None:
                print(f"  WARNING: unknown IO code '{io_code}' for industry '{company_industry}'", file=sys.stderr)
                continue
            db.add(IoIndustryMapping(
                company_industry=company_industry,
                io_sector_id=io_sector_id,
                confidence=confidence,
                note=note or None,
            ))
            mapping_count += 1
        db.commit()
        print(f"  inserted {mapping_count} industry mappings")

    print("done")


def _to_float(val: object) -> float | None:
    if val is None:
        return None
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def _compute_direct_consumption_coefficients(
    db,
    IoFlowMatrix,  # noqa: N803
    IoSector,      # noqa: N803
    sector_id_map: dict[int, int],
) -> None:
    """Compute a_ij = X_ij / X_j and update io_flow_matrix rows."""
    # load total_input for each sector (col j's total_input)
    sector_total_input: dict[int, float] = {}
    for sector in db.query(IoSector).all():
        if sector.total_input:
            sector_total_input[sector.id] = sector.total_input

    updated = 0
    for row in db.query(IoFlowMatrix).all():
        xj = sector_total_input.get(row.col_sector_id)
        if xj and xj != 0:
            row.direct_consumption_coeff = row.flow_value / xj
            updated += 1
    db.commit()
    print(f"  computed {updated} direct consumption coefficients")


if __name__ == "__main__":
    main()
